import re
import openpyxl
import xmind
from xmind.core.markerref import MarkerId

class Excel2Xmind():

    def design_sheet(self,dicts,xmind_file):
        workbook = xmind.load(xmind_file)
        sheet = workbook.getPrimarySheet()
        sheet.setTitle('First Sheet')
        root_topic = sheet.getRootTopic()
        root_topic.setTitle('root node')
        self.dict_item(dicts,root_topic)
        xmind.save(workbook)

    def dict_item(self,dicts,topic):
        for key in dicts:
            split_result = key.split('$')
            if len(split_result) == 1:
                subtopic = topic.addSubTopic()
                subtopic.setTitle(split_result[0])
            elif len(split_result) == 2:
                subtopic = topic.addSubTopic()
                subtopic.setTitle(split_result[1])
                self.priority_mark(split_result[0],subtopic)
            elif len(split_result) == 3:
                subtopic = topic.addSubTopic()
                subtopic.setTitle(split_result[1])
                self.priority_mark(split_result[0],subtopic)
                subtopic.setPlainNotes(split_result[2])
            else:
                raise KeyError

            if isinstance(dicts[key],dict):
                self.dict_item(dicts[key],subtopic)
            if isinstance(dicts[key],str):
                subtopic = subtopic.addSubTopic()
                subtopic.setTitle(dicts[key])
        return True

    def priority_mark(self,str,subtopic):
        if str == 'P1':
            return subtopic.addMarker(MarkerId.priority1)
        elif str == 'P2':
            return subtopic.addMarker(MarkerId.priority2)
        elif str == 'P3':
            return subtopic.addMarker(MarkerId.priority3)
        elif str == 'P4':
            return subtopic.addMarker(MarkerId.priority4)
        elif str == 'P5':
            return subtopic.addMarker(MarkerId.priority5)
        elif str == 'P6':
            return subtopic.addMarker(MarkerId.priority6)
        elif str == 'P7':
            return subtopic.addMarker(MarkerId.priority7)
        else :
            return subtopic.addMarker(MarkerId.priority8)

    def load_excel(self,filename):
        wb= openpyxl.load_workbook(filename)
        ws = wb.active

        data = []
        for row in range(2,ws.max_row+1):
            if ws[row][4].value is None:
                raise RuntimeError('某个用例缺少测试步骤哦！')

            if ws[row][3].value not in [None,'']:
                case = ws[row][2].value+"$"+ws[row][1].value+"$"+ws[row][3].value
            else:
                case = ws[row][2].value+"$"+ws[row][1].value

            if ws[row][6].value in [None,'']:
                pre_result = "cases" + "|" + case
            else:
                pre_result = ws[row][6].value+"|"+"cases"+"|"+case

            if ws[row][5].value != None:
                miaoshu = ws[row][4].value.split('\n')
                jieguo = ws[row][5].value.split('\n')

                if len(miaoshu) < len(jieguo):
                    raise RuntimeError('有用例结果比步骤还多哦！')

                for i in range(len(miaoshu)):
                    if i<len(jieguo):
                        miaoshu[i] = re.sub('^\d*\.','',miaoshu[i])
                        jieguo[i] = re.sub('^\d*\.','',jieguo[i])
                        result = (pre_result+"|"+miaoshu[i]+"|"+jieguo[i])
                    else:
                        miaoshu[i] = re.sub('^\d*\.', '',miaoshu[i])
                        result = (pre_result+"|"+miaoshu[i]+"|"+'')

                    after = result.split("|")
                    final = self.list2dict(after)
                    data.append(final)
                    print(data)
                #after = result.split("$")
                #print(hehe(after)
                # data.append(hehe(after))
            else:
                miaoshu = ws[row][4].value.split('\n')
                for i in range(len(miaoshu)):
                    miaoshu[i] = re.sub('^\d*\.', '', miaoshu[i])
                    result = (pre_result + "|" + miaoshu[i])
                    after = result.split("|")
                    final = self.list2dict(after)
                    data.append(final)
                # ws[row][4].value = re.sub('^\d*\.','',ws[row][4].value)
                # ws[row][5].value = re.sub('^\d*\.', '', ws[row][5].value)
                # result= pre_result+"|"+ws[row][4].value+"|"+ws[row][5].value
                # after = result.split("|")
                # final = self.list2dict(after)
                # data.append(final)
            #print(data)

        result = self.combine_dict(data)
        return result

    def combine_dict(self,l1):
        o = {}
        for d in l1:
            n = o
            m = d
            p = None
            while isinstance(m, dict):
                if p is not None:
                    if k not in p:
                        p[k] = {}
                    n = p[k]
                (k, m), = m.items()
                p = n

            while isinstance(m,str):
                p.setdefault(k,m)
                break
        return o

    def list2dict(self,list_):
        dict_={}
        if len(list_)>2:
            dict_[list_[0]]=self.list2dict(list_[1:])
        else:
            dict_[list_[0]]=list_[1]
        return dict_

if __name__ == '__main__':
    final = Excel2Xmind()
    finaldata = final.load_excel('test.xlsx')
    final.design_sheet(finaldata,'test.xmind')









