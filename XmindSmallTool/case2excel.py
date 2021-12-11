from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.styles import Border, Side, Font, Alignment
from config import ConfigParser


class Case2Excel:
    def __init__(self, tpath, npath):
        self.template = load_workbook(tpath).active
        self.case_excel = npath

    def get_template_field(self):
        # 获取模板中的关键字段所在的行列
        field_config = ConfigParser.get_config('field', 'template')
        field_number = { key:'' for key in field_config.keys() }
        for row in self.template.iter_rows():
            for cell in row:
                if cell.value:
                    for k,v in field_config.items():
                        if v in cell.value:
                            field_number[k] = cell.column
        field_number['row'] = self.template.max_row

        return field_number

    def copy_template_to_new(self, src, tag):

        max_row = src.max_row  # 最大行数
        max_column = src.max_column  # 最大列数

        wm = list(src.merged_cells) # 开始处理合并单元格
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell2 = str(wm[i]).replace('(<MergeCell ', '').replace('>,)', '')
                # print("MergeCell : %s" % cell2)
                tag.merge_cells(cell2)

        for m in range(1, max_row + 1):
            tag.row_dimensions[m].height = src.row_dimensions[m].height
            for n in range(1, 1 + max_column):
                if n < 27:
                    c = chr(n + 64).upper()  # ASCII字符,chr(65)='A'
                else:
                    if n < 677:
                        c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
                    else:
                        c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(
                            divmod(divmod(n, 676)[1], 26)[1] + 64)
                i = '%s%d' % (c, m)  # 单元格编号
                if m == 1:
                    #				 print("Modify column %s width from %d to %d" % (n, ws2.column_dimensions[c].width ,ws.column_dimensions[c].width))
                    tag.column_dimensions[c].width = src.column_dimensions[c].width
                try:
                    # getattr(src.cell(row=m, column=c), "value")
                    cell1 = src[i]  # 获取data单元格数据
                    tag[i].value = cell1.value  # 赋值到ws2单元格
                    if cell1.has_style:  # 拷贝格式
                        tag[i].font = copy(cell1.font)
                        tag[i].border = copy(cell1.border)
                        tag[i].fill = copy(cell1.fill)
                        tag[i].number_format = copy(cell1.number_format)
                        tag[i].protection = copy(cell1.protection)
                        tag[i].alignment = copy(cell1.alignment)
                except AttributeError as e:
                    # print("cell(%s) is %s" % (i, e))
                    continue
        return tag

    def write_case_to_excel(self, data):
        wb = Workbook()
        field = self.get_template_field()
        # print(field)

        wb.remove_sheet(wb.active)  # 删除默认的 Sheet1

        for d in data:
            print(d)
            new_sheet = wb.create_sheet(d['map_name'])
            new_sheet = self.copy_template_to_new(self.template, new_sheet)
            start = new_sheet.max_row + 1
            for test_case in d['test_cases']:
                # page_name = test_case.get('page')
                # function_name = test_case.get('登录')

                row = new_sheet.max_row + 1
                new_sheet.cell(row, field['module'], value=test_case['module'])
                new_sheet.cell(row, field['priority'], value=test_case['priority'])
                new_sheet.cell(row, field['point'], value=test_case['point'] or '')
                new_sheet.cell(row, field['test_title'], value=test_case['test_title'])
                new_sheet.cell(row, field['test_step'], value=test_case['test_steps'] or '')
                new_sheet.cell(row, field['test_result'], value=test_case['test_results'] or '')
                new_sheet.cell(row, field['preset'], value=test_case['preset'] or '')
                if field['smoke']:
                    new_sheet.cell(row, field['smoke'], value=test_case['is_smoke'] or '')
                if field['category']:
                    new_sheet.cell(row, field['category'], value=test_case['category'] or '')
                if field['stage']:
                    new_sheet.cell(row, field['stage'], value=test_case['stage'] or '')

            # 处理样式
            for row in new_sheet.iter_rows(min_row=start):
                for cell in row:
                    # 设置边框
                    thin = Side(border_style="thin", color="000000")
                    border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.border = border

                    # 自动换行
                    cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

                    # 设置字体
                    cell.font = Font(name='微软雅黑', size=ConfigParser.get_config('font_size', 'style'))

        wb.save(self.case_excel)



if __name__ == '__main__':
    from parsercases import ParserCases

    xc = ParserCases('../test/test.xmind')
    ce = Case2Excel('template.xlsx', '../test/1.xlsx')
    ce.write_case_to_excel(xc.all_map_case)
    # for sheet in xc.all_map_case:
    #     print(sheet)
    #     ce.write_case_to_excel(sheet)